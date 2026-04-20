"""XLSX adapter — one ``Unit`` per non-empty text cell.

Workbook structure (sheets, column widths, formulas, formatting) is
preserved by loading the source workbook on write and mutating only the
cells we translated. Formula cells are deliberately skipped — we never
rewrite expressions.
"""

from __future__ import annotations

from collections.abc import Iterable
from pathlib import Path

from openpyxl import load_workbook

from ...domain import TranslatedUnit, Unit, UnitKind
from ...use_cases.ports import DocumentAdapter


class XlsxAdapter(DocumentAdapter):
    extension = ".xlsx"

    def extract(self, source_path: Path) -> list[Unit]:
        workbook = load_workbook(source_path, data_only=False, read_only=False)
        units: list[Unit] = []
        idx = 0
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type in ("f", "n", "d", "b"):
                        continue
                    value = cell.value
                    if value is None:
                        continue
                    if not isinstance(value, str):
                        continue
                    stripped = value.strip()
                    if not stripped:
                        continue
                    units.append(
                        Unit(
                            id=f"{idx:04d}",
                            kind=UnitKind.CELL,
                            text=value,
                            meta={
                                "sheet": sheet.title,
                                "coord": cell.coordinate,
                            },
                        )
                    )
                    idx += 1
        workbook.close()
        return units

    def write(
        self,
        source_path: Path,
        translated: Iterable[TranslatedUnit],
        target_lang: str,
        output_path: Path,
    ) -> None:
        workbook = load_workbook(source_path, data_only=False, read_only=False)
        try:
            for unit in translated:
                sheet_name = unit.meta.get("sheet")
                coord = unit.meta.get("coord")
                if not sheet_name or not coord:
                    continue
                if sheet_name not in workbook.sheetnames:
                    continue
                workbook[sheet_name][coord] = unit.target_text
            workbook.save(output_path)
        finally:
            workbook.close()
